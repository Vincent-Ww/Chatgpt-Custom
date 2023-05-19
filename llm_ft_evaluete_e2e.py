# -- coding: utf-8 --
# @Author : wuzixun
# @Time : 2023/5/19 11:26 AM
# @File : llm_ft_evaluete_e2e.py

import torch
from transformers import AutoTokenizer, AutoModel
from peft import get_peft_model, LoraConfig, TaskType
from openpyxl import Workbook
import json
import requests
from tqdm import tqdm
import pandas as pd

PEFT_PATH = "./output-20230518/adapter_model.bin"
CHATGLM_PATH = "/home/xiezizhe/wuzixun/LLM/chatglm-6b"
DEV_DATA_PATH = "/home/xiezizhe/wuzixun/LLM/ChatGLM-Tuning/data/ks_h2h_combine_2w/resample_instruce_dialogue_combine_dev.json"


def chatglm_inference(model, tokenizer, dialogue):
    instruction = """下面是用户和客服的一段对话，请你根据下面这段对话总结用户的诉求和FT。
        FT是用户描述用户诉求的多层的树状知识结构，不同层级之间用-相连。(e.g. 电商-买家, 平台-账号)

        请按照以下格式：
        诉求：<在这里输出用户的诉求>
        FT: <在这里输出FT>

        如果用户没有诉求，输出[其他问题]，字数限制在50个字以内。如果用户没有FT，则输出[无]。"""
    context = f"Instruction: {instruction}\n"
    context += f"Input: {dialogue}\n"
    context += "Answer: "

    response, _ = model.chat(tokenizer, context, history=[])
    return response


def call_faq_rpc(text):
    url = "https://kess.corp.kuaishou.com/api/fe/v2/central/debugger/v1/call"
    cookie = "_did=web_26641810720AE322; ksCorpDeviceid=b686bc1f-2c22-4935-a198-db17a8bdce20; " \
             "OUTFOX_SEARCH_USER_ID_NCOO=1072921961.3413543; clientid=3; did=web_61a5a7e54f0bb8c230d66cae6f2525bc; " \
             "Hm_lvt_86a27b7db2c5c0ae37fee4a8a35033ee=1663123444; client_key=65890b29; " \
             "adm-did=iYJlifMR62nx9UjJVcyjnqW4W0GLYIv0EDxA229r7KA5QA7VUQAiJHgFBjqY7CiR; " \
             "_ga=GA1.2.1604015848.1658303187; _ga_17SWBP5QSM=GS1.1.1668758690.2.0.1668758692.0.0.0; " \
             "hdige2wqwoino=TQDrbrjWNkHhCys3czwCwbaCFC2BPKDYfc874b1b; " \
             "apdid=86ec03ea-00b5-4037-9511-f74a06962e411df5df9c7d42f54ea46155cfbde73ac2:1678676019:1; " \
             "didv=1678678973000; ajs_user_id=4ebd0208-8328-5d69-8c44-ec50939c0967; " \
             "aidp_cookie_key=1f7CwEdi/1nUe+WlWZmFvPGJqT0wbdGxUGrxhziXVPQ9Czr0x4r7vBM64lCTshlgjIAiXLTsW0FW" \
             "/OB9U3w1mGULR+MpUpYmeL7lLGr/N6ZObThsO9aKib4fq11xdvsrLgjH5zI0eCGTpCeNevs/OqimKC7YkNqHVts" \
             "+wVUZkNhU5b6rb4UTIOFNn+nCFUOpPHis4myRJKuE31oZ3SWI+g==; " \
             "ajs_anonymous_id=9d85dba7-0cbd-4343-a0ca-f86b4b0570cd; " \
             "KSPP=R42vs6jkIEx5lDfXn5gz98gLofw1o8VS+1b1zddPVZSGDDgqjFvf1w==; " \
             "JSESSIONID=05C81CBF9C8AD07AD10D2960E175B184 "
    retry_time = 5
    while retry_time > 0:
        try:
            response = requests.post(
                url=f"{url}",
                headers={
                    "Content-Type": "application/json",
                    "Cookie": f"{cookie}",
                    "origin": "https://kess.corp.kuaishou.com",
                    "referer": "https://kess.corp.kuaishou.com/",
                    "authority": "kess.corp.kuaishou.com"
                },
                data=json.dumps({
                    "hostInfo": "10.28.141.141:22852",
                    "kessName": "grpc_FaqRecognizeRpcService",
                    "laneId": "",
                    "methodName": "kuaishou.admin.csc.FaqRecognitionRpcService/Recognize",
                    "request": faq_request(text)
                })

            )
            res = json.loads(response.json()["response"])
            return res
        except Exception as e:
            print(e, "failed to call the nlu api")
            retry_time -= 1
        finally:
            if retry_time == 0:
                return ""


def faq_request(text):
    request = json.dumps({
        "agent_id": "d7188f92-75a5-46d8-bade-3200c26e1065",
        "user_id": "1234",
        "text": text,
        "op_key": "{\"botStraightAnswerThreshold\": 0.986,        \"botClarifyAnswerThreshold\": 0.95,        "
                  "\"botId\": \"657feaec-e299-4e7e-875b-e8d9c4676643\",        \"ftScore\": {},        "
                  "\"confirmSize\": 5,        \"userId\": 99    }",
        "timeout_ms": "0"
    })
    return request


def get_faq_id_name_map(file_path):
    with open(file_path, "r") as f:
        data = json.load(f)
    faqs = data['data']['list']
    faq_id2name = {}
    faq_name2id = {}
    for ele in faqs:
        id = ele['id']
        name = ele['question']
        faq_id2name[id] = name
        faq_name2id[name] = id
    return faq_name2id, faq_id2name


def get_faq_ft_map(file_path):
    data = pd.read_excel(file_path, sheet_name='标准问')
    faq2ft = {}
    ft2faq = {}
    for i in range(data.shape[0]):
        faq = data.iloc[i]['标准问名称']
        multi_ft = data.iloc[i]['类目信息'].split("/")
        ft1, ft2 = "", ""
        if len(multi_ft) >= 3:
            ft1 = multi_ft[2]
        if len(multi_ft) >= 4:
            ft2 = multi_ft[3]
        ft_1_2 = ft1 + "-" + ft2
        faq2ft[faq] = ft_1_2
        ft2faq[ft_1_2] = faq
    return faq2ft, ft2faq


if __name__ == "__main__":

    tokenizer = AutoTokenizer.from_pretrained(CHATGLM_PATH, trust_remote_code=True)
    model = AutoModel.from_pretrained(CHATGLM_PATH, load_in_8bit=True, trust_remote_code=True, device_map='auto')
    model = model.eval()

    # 这里的配置需要跟自己的LoRA一样
    peft_config = LoraConfig(
        task_type=TaskType.CAUSAL_LM,
        inference_mode=False,
        r=8,
        lora_alpha=32,
        lora_dropout=0.1,
        target_modules=['dense', 'dense_h_to_4h', 'dense_4h_to_h', 'query_key_value']
    )

    model_lora = get_peft_model(model, peft_config)
    model_lora.load_state_dict(torch.load(PEFT_PATH), strict=False)
    torch.set_default_tensor_type(torch.cuda.FloatTensor)

    with open(DEV_DATA_PATH, "r", encoding='utf-8') as f:
        dev_data = json.load(f)

    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(1, 1).value = '对话'
    sheet.cell(1, 2).value = 'LLM结果(微调后)'
    sheet.cell(1, 3).value = '标准问'
    sheet.cell(1, 4).value = 'FT'
    sheet.cell(1, 5).value = "FAQ识别后对应FT"
    sheet.cell(1, 6).value = "一级FT是否正确"
    sheet.cell(1, 7).value = "二级FT是否正确"

    _, faq_id2name = get_faq_id_name_map("/home/xiezizhe/wuzixun/Chatgpt-Custom/ChatGLM-LoRA/data/response.json")
    faq2ft, _ = get_faq_ft_map("/home/xiezizhe/wuzixun/Chatgpt-Custom/ChatGLM-LoRA/data/IVR_FAQ.xlsx")

    nrow = 2
    for sample in tqdm(dev_data):
        try:
            sheet.cell(nrow, 1).value = sample['input']
            sheet.cell(nrow, 2).value = chatglm_inference(model, tokenizer, sample['input'])
            sheet.cell(nrow, 3).value = sample['question']

            # sheet.cell(nrow, 4).value = sample['ft_name']
            multi_ft_label = sample['FT'].split("-")
            if len(multi_ft_label) >= 1:
                ft_1_label = multi_ft_label[0]
            if len(multi_ft_label) >= 2:
                ft_2_label = multi_ft_label[1]
            ft_1_2_label = ft_1_label + "-" + ft_2_label
            sheet.cell(nrow, 4).value = ft_1_2_label

            faq_output = call_faq_rpc(sample['question'])
            faq_list = faq_output['items']
            if len(faq_list) > 0:
                faq_id = faq_list[0]['question_id']
                faq_name = faq_id2name[faq_id]
                ft = faq2ft[faq_name]
            else:
                ft = ""
            sheet.cell(nrow, 5).value = ft


            def is_first_equal(ft_label, ft_predict):
                label = ft_label.split("-")[0]
                predict = ft_predict.split("-")[0]
                return label == predict


            def is_second_equal(ft_label, ft_predict):
                label = ft_label.split("-")
                predict = ft_predict.split("-")
                return label == predict


            sheet.cell(nrow, 6).value = is_first_equal(ft_1_2_label, ft)
            sheet.cell(nrow, 7).value = is_second_equal(ft_1_2_label, ft)

            nrow += 1
            if nrow % 100 == 0:
                workbook.save(f"人人对话LLM验证集{nrow}(faq识别后)_com.xlsx")
                print(f"人人对话LLM验证集{nrow}(faq识别后)_com.xlsx: Save!")
        except Exception as e:
            print("Exception: ", e)
    workbook.save(f"人人对话LLM验证集{nrow}(faq识别后)_com.xlsx")
