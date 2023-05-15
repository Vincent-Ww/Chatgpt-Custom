import sys
import torch
import json
from transformers import AutoTokenizer, AutoModel
from peft import get_peft_model, LoraConfig, TaskType
import pandas
from tqdm import tqdm
from openpyxl import Workbook


def chatglm_inference(model, tokenizer, dialogue):
    prompt = "从用户和客服的对话中提炼用户的诉求问题:\n" + dialogue
    response, _ = model.chat(tokenizer, prompt, history=[])
    return response


if __name__ == "__main__":
    if len(sys.argv) == 2:
        chatglm_path = sys.argv[1]
    else:
        chatglm_path = "/home/xiezizhe/wuzixun/LLM/chatglm-6b"
        dev_data_path = "/home/xiezizhe/wuzixun/LLM/Chatgpt-Custom/ChatGLM-LoRA/data/ks_human2human_question/resample_instruce_dialogue_dev.json"

    tokenizer = AutoTokenizer.from_pretrained(chatglm_path, trust_remote_code=True)
    model = AutoModel.from_pretrained(chatglm_path, load_in_8bit=True, trust_remote_code=True, device_map='auto')
    model = model.eval()

    response, history = model.chat(tokenizer, '你是谁', history=[])

    peft_path = "./output-20230425/adapter_model.bin"
    # 这里的配置需要跟自己的LoRA一样
    peft_config = LoraConfig(
        task_type=TaskType.CAUSAL_LM,
        inference_mode=False,
        r=8,
        lora_alpha=32,
        lora_dropout=0.1,
        target_modules=['dense', 'dense_h_to_4h', 'dense_4h_to_h', 'query_key_value']
    )

    model_lora = get_peft_model(model, peft_config)
    model_lora.load_state_dict(torch.load(peft_path), strict=False)
    torch.set_default_tensor_type(torch.cuda.FloatTensor)

    with open(dev_data_path, "r", encoding='utf-8') as f:
        dev_data = json.load(f)

    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(1, 1).value = '对话'
    sheet.cell(1, 2).value = 'LLM结果(微调后)'
    sheet.cell(1, 3).value = '标准问'
    sheet.cell(1, 4).value = 'FT'

    nrow = 2
    for sample in tqdm(dev_data):
        try:
            sheet.cell(nrow, 1).value = sample['input']
            sheet.cell(nrow, 2).value = chatglm_inference(model, tokenizer, sample['input'])
            sheet.cell(nrow, 3).value = sample['question']
            sheet.cell(nrow, 4).value = sample['ft_name']
            nrow += 1
            if nrow % 20 == 0:
                workbook.save(f"人人对话LLM验证集{nrow}.xlsx")
                print(f"人人对话LLM验证集{nrow}.xlsx: Save!")
        except Exception as e:
            print("Exception: ", e)
    workbook.save(f"人人对话LLM验证集 {nrow}.xlsx")
